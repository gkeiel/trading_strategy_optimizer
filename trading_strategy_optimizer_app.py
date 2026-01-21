import os, sys, json, threading
from openpyxl import load_workbook
from trading_strategy_optimizer import run_tso
from core_app.gui import Gui
from core_app.redirector import Redirector
os.chdir(os.path.dirname(os.path.abspath(__file__)))


def run_tso_app():
    app = None
    
    def on_log(msg):
        app.append_log(msg)
        
    def load_metrics():
        wb = load_workbook("data/results/results.xlsx", data_only=True)
        ws = wb.active
        
        headers = [cell.value for cell in ws[1]]
        values  = [cell.value for cell in ws[2]]
        results = dict(zip(headers, values))
        print("\n======= METRICS =======")
        #print(f"Params                  {results['Parameters']}")
        print(f"Score                       {float(results['Score']):.2f}")
        print(f"Strategy return    {float(results['Return_Strategy'])*100:.2f} %")
        print(f"Trades                    {results['Trades']}")
        print(f"Sharpe                    {float(results['Sharpe']):.2f}")
        print(f"Max Drawdown   {float(results['Max_Drawdown'])*100:.2f} %")
        print(f"Market return       {float(results['Return_Market'])*100:.2f} %")
        print("======================\n")
        
    def run_thread(config):
        with open("config/config.json", "w") as f: json.dump(config, f, indent=4)
        stdout_original = sys.stdout
        stderr_original = sys.stderr

        try:
            sys.stdout = Redirector(on_log)
            sys.stderr = Redirector(on_log)
            run_tso(on_log=on_log)
            on_log("Optimization completed.")
            load_metrics()

        except Exception as e:
            on_log(f"Error: {e}")

        finally:
            sys.stdout = stdout_original
            sys.stderr = stderr_original

    def on_start(config):
        threading.Thread(target=run_thread, args=(config,), daemon=True).start()
    
    app = Gui(on_start_callback=on_start)
    app.run()
        

if __name__ == "__main__":
    run_tso_app()
